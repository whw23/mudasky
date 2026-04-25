"""excel_io 工具单元测试。

测试 Excel/ZIP 读写和占位图片生成。
"""

import io
import zipfile

from openpyxl import Workbook

from app.utils.excel_io import (
    create_placeholder_image,
    create_zip,
    extract_zip,
    load_workbook_from_bytes,
    read_sheet_rows,
    workbook_to_bytes,
    write_sheet_header,
)


# ---- extract_zip ----


class TestExtractZip:
    """ZIP 解压测试。"""

    def test_normal_files(self):
        """正常解压 ZIP 文件。"""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("file1.txt", "hello")
            zf.writestr("file2.txt", "world")
        result = extract_zip(buf.getvalue())
        assert result == {
            "file1.txt": b"hello",
            "file2.txt": b"world",
        }

    def test_skip_macos_hidden(self):
        """跳过 macOS 隐藏文件（__ 和 . 开头）。"""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("normal.txt", "ok")
            zf.writestr("__MACOSX/file", "skip")
            zf.writestr(".hidden", "skip")
        result = extract_zip(buf.getvalue())
        assert list(result.keys()) == ["normal.txt"]

    def test_empty_zip(self):
        """空 ZIP 返回空字典。"""
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w"):
            pass
        result = extract_zip(buf.getvalue())
        assert result == {}


# ---- create_zip ----


class TestCreateZip:
    """ZIP 打包测试。"""

    def test_creates_valid_zip(self):
        """创建有效 ZIP。"""
        files = {"a.txt": b"content_a", "b.txt": b"content_b"}
        result = create_zip(files)

        with zipfile.ZipFile(io.BytesIO(result)) as zf:
            assert set(zf.namelist()) == {"a.txt", "b.txt"}
            assert zf.read("a.txt") == b"content_a"

    def test_empty_files(self):
        """空文件字典创建空 ZIP。"""
        result = create_zip({})
        with zipfile.ZipFile(io.BytesIO(result)) as zf:
            assert zf.namelist() == []


# ---- read_sheet_rows / write_sheet_header ----


class TestSheetOperations:
    """Sheet 读写测试。"""

    def test_read_with_skip_header(self):
        """跳过表头读取行。"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", 25])

        rows = read_sheet_rows(ws, skip_header=True)
        assert len(rows) == 2
        assert rows[0] == ["Alice", 30]

    def test_read_without_skip_header(self):
        """不跳过表头读取行。"""
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Age"])
        ws.append(["Alice", 30])

        rows = read_sheet_rows(ws, skip_header=False)
        assert len(rows) == 2
        assert rows[0] == ["Name", "Age"]

    def test_write_header(self):
        """写入表头。"""
        wb = Workbook()
        ws = wb.active
        write_sheet_header(ws, ["ID", "Name", "Status"])

        assert ws.cell(1, 1).value == "ID"
        assert ws.cell(1, 2).value == "Name"
        assert ws.cell(1, 3).value == "Status"


# ---- workbook_to_bytes / load_workbook_from_bytes ----


class TestWorkbookSerialization:
    """Workbook 序列化测试。"""

    def test_roundtrip(self):
        """序列化后反序列化保留数据。"""
        wb = Workbook()
        ws = wb.active
        ws.append(["test", 123])

        data = workbook_to_bytes(wb)
        wb2 = load_workbook_from_bytes(data)

        assert wb2.active.cell(1, 1).value == "test"
        assert wb2.active.cell(1, 2).value == 123


# ---- create_placeholder_image ----


class TestCreatePlaceholderImage:
    """占位图片生成测试。"""

    def test_returns_png_bytes(self):
        """返回有效 PNG 字节。"""
        data = create_placeholder_image()
        # PNG 魔数
        assert data[:4] == b"\x89PNG"

    def test_custom_size(self):
        """自定义尺寸生成。"""
        data = create_placeholder_image(width=100, height=100)
        assert len(data) > 0
