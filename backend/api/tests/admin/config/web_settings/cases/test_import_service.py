"""ImportService（成功案例）单元测试 — generate_template / _parse_workbook。

测试成功案例批量导入的模板生成和 Excel 解析逻辑。
"""

import io
import zipfile
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import Workbook, load_workbook

from api.admin.config.web_settings.cases.import_service import (
    ImportService,
)

from .conftest_import import (
    CASE_REPO,
    UNI_REPO,
    create_valid_workbook,
    make_case,
    make_university,
)


@pytest.fixture
def service(mock_session) -> ImportService:
    """构建 ImportService 实例。"""
    return ImportService(mock_session)


# ---- generate_template ----


def test_generate_template_returns_valid_zip(service):
    """生成导入模板返回有效的 ZIP 文件。"""
    result = service.generate_template()

    assert isinstance(result, bytes)
    assert len(result) > 0

    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        names = zf.namelist()
        assert any(n.endswith(".xlsx") for n in names)
        assert any(n.startswith("images/") for n in names)


def test_generate_template_has_required_headers(service):
    """生成的模板包含必填字段列头。"""
    result = service.generate_template()

    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        xlsx = [n for n in zf.namelist() if n.endswith(".xlsx")][0]
        wb = load_workbook(io.BytesIO(zf.read(xlsx)))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "学生姓名" in headers
        assert "院校名称" in headers
        assert "专业" in headers
        assert "入学年份" in headers


def test_generate_template_has_sample_row(service):
    """生成的模板包含示例数据行。"""
    result = service.generate_template()

    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        xlsx = [n for n in zf.namelist() if n.endswith(".xlsx")][0]
        wb = load_workbook(io.BytesIO(zf.read(xlsx)))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 1
        assert rows[0][0] == "张三"


# ---- _parse_workbook ----


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(CASE_REPO)
async def test_parse_workbook_new_case(
    mock_repo, mock_uni_repo, service
):
    """解析有效 workbook，新案例返回 status=new。"""
    wb = create_valid_workbook()
    mock_repo.find_case = AsyncMock(return_value=None)
    mock_uni_repo.get_university_by_name = AsyncMock(
        return_value=None
    )

    items, errors = await service._parse_workbook(wb)

    assert len(items) == 1
    assert items[0]["student_name"] == "张三"
    assert items[0]["status"] == "new"
    assert items[0]["data"]["year"] == 2026
    assert len(errors) == 0


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(CASE_REPO)
async def test_parse_workbook_existing_update(
    mock_repo, mock_uni_repo, service
):
    """已存在且有变化的案例返回 status=update。"""
    wb = create_valid_workbook()
    existing = make_case(program="物理学", is_featured=False)
    mock_repo.find_case = AsyncMock(return_value=existing)
    mock_uni_repo.get_university_by_name = AsyncMock(
        return_value=None
    )

    items, errors = await service._parse_workbook(wb)

    assert len(items) == 1
    assert items[0]["status"] == "update"
    assert "program" in items[0]["changed_fields"]


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(CASE_REPO)
async def test_parse_workbook_with_university_match(
    mock_repo, mock_uni_repo, service
):
    """解析时匹配到已有院校则设置 university_id。"""
    wb = create_valid_workbook()
    mock_repo.find_case = AsyncMock(return_value=None)
    mock_uni_repo.get_university_by_name = AsyncMock(
        return_value=make_university()
    )

    items, errors = await service._parse_workbook(wb)

    assert items[0]["data"]["university_id"] == "uni-1"


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(CASE_REPO)
async def test_parse_workbook_empty_row_skipped(
    mock_repo, mock_uni_repo, service
):
    """空行被跳过。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成功案例"
    ws.append(["学生姓名", "院校名称", "专业", "入学年份"])
    ws.append([None, None, None, None])

    items, errors = await service._parse_workbook(wb)

    assert len(items) == 0
    assert len(errors) == 0


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(CASE_REPO)
async def test_parse_workbook_missing_university(
    mock_repo, mock_uni_repo, service
):
    """缺少院校名称报错。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成功案例"
    _h = ["学生姓名", "院校", "专业", "年份", "感言", "头像", "Offer", "精选", "排序"]
    ws.append(_h)
    ws.append(["张三", None, "CS", 2026, None, None, None, None, None])

    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)

    items, errors = await service._parse_workbook(wb)

    assert len(errors) == 1
    assert "院校名称" in errors[0]["error"]


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(CASE_REPO)
async def test_parse_workbook_missing_program(
    mock_repo, mock_uni_repo, service
):
    """缺少专业报错。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成功案例"
    _h = ["学生姓名", "院校", "专业", "年份", "感言", "头像", "Offer", "精选", "排序"]
    ws.append(_h)
    ws.append(["张三", "哈佛", None, 2026, None, None, None, None, None])

    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)

    items, errors = await service._parse_workbook(wb)

    assert len(errors) == 1
    assert "专业" in errors[0]["error"]


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(CASE_REPO)
async def test_parse_workbook_missing_year(
    mock_repo, mock_uni_repo, service
):
    """缺少入学年份报错。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成功案例"
    _h = ["学生姓名", "院校", "专业", "年份", "感言", "头像", "Offer", "精选", "排序"]
    ws.append(_h)
    ws.append(["张三", "哈佛", "CS", None, None, None, None, None, None])

    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)

    items, errors = await service._parse_workbook(wb)

    assert len(errors) == 1
    assert "入学年份" in errors[0]["error"]


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(CASE_REPO)
async def test_parse_workbook_invalid_year(
    mock_repo, mock_uni_repo, service
):
    """入学年份不是整数报错。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "成功案例"
    _h = ["学生姓名", "院校", "专业", "年份", "感言", "头像", "Offer", "精选", "排序"]
    ws.append(_h)
    ws.append(["张三", "哈佛", "CS", "abc", None, None, None, None, None])

    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)

    items, errors = await service._parse_workbook(wb)

    assert len(errors) == 1
    assert "入学年份" in errors[0]["error"]


# ---- _guess_mime_type ----


def test_guess_mime_type_jpg(service):
    """jpg 返回 image/jpeg。"""
    assert service._guess_mime_type("photo.jpg") == "image/jpeg"


def test_guess_mime_type_png(service):
    """png 返回 image/png。"""
    assert service._guess_mime_type("icon.png") == "image/png"


def test_guess_mime_type_unknown(service):
    """未知扩展名返回 image/jpeg。"""
    assert service._guess_mime_type("file.bmp") == "image/jpeg"


def test_guess_mime_type_webp(service):
    """webp 返回 image/webp。"""
    assert service._guess_mime_type("photo.webp") == "image/webp"
