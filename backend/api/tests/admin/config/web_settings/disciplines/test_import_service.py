"""DisciplineImportService 单元测试 - 模板生成与预览。

测试学科分类批量导入的 generate_template 和 preview 方法。
"""

import io
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook, load_workbook

from api.admin.config.web_settings.disciplines.import_service import (
    DisciplineImportService,
)
from app.db.discipline.models import Discipline, DisciplineCategory

DISC_REPO = (
    "api.admin.config.web_settings.disciplines.import_service.disc_repo"
)


def _make_category(
    category_id: str = "cat-1", name: str = "工学"
) -> MagicMock:
    """创建模拟 DisciplineCategory 对象。"""
    cat = MagicMock(spec=DisciplineCategory)
    cat.id = category_id
    cat.name = name
    return cat


def _make_discipline(
    discipline_id: str = "disc-1", name: str = "计算机科学"
) -> MagicMock:
    """创建模拟 Discipline 对象。"""
    disc = MagicMock(spec=Discipline)
    disc.id = discipline_id
    disc.name = name
    return disc


def _create_workbook(rows: list[list]) -> bytes:
    """创建包含指定行数据的 Excel 字节。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "学科分类"
    ws.append(["大分类", "小分类"])
    for row in rows:
        ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


@pytest.fixture
def service(mock_session) -> DisciplineImportService:
    """构建 DisciplineImportService 实例。"""
    return DisciplineImportService(mock_session)


# ---- generate_template ----


class TestGenerateTemplate:
    """generate_template 模板生成测试。"""

    def test_returns_valid_excel_bytes(self, service):
        """生成模板返回有效的 Excel 字节。"""
        result = service.generate_template()
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_template_has_header_and_examples(self, service):
        """模板包含表头和示例数据。"""
        result = service.generate_template()
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.title == "学科分类"
        headers = [cell.value for cell in ws[1]]
        assert "大分类" in headers
        assert "小分类" in headers
        # 至少有 4 行示例
        data_rows = list(
            ws.iter_rows(min_row=2, values_only=True)
        )
        assert len(data_rows) >= 4

    def test_template_example_content(self, service):
        """模板示例行包含正确的学科数据。"""
        result = service.generate_template()
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active
        row2 = [cell.value for cell in ws[2]]
        assert row2[0] == "工学"
        assert row2[1] == "计算机科学"


# ---- preview ----


class TestPreview:
    """preview 预览测试。"""

    @pytest.mark.asyncio
    @patch(DISC_REPO)
    async def test_new_category_and_discipline(
        self, mock_repo, service
    ):
        """大分类和小分类都不存在，status 为 new。"""
        content = _create_workbook([["工学", "计算机科学"]])
        mock_repo.get_category_by_name = AsyncMock(
            return_value=None
        )

        result = await service.preview(content)

        assert len(result["items"]) == 1
        item = result["items"][0]
        assert item["status"] == "new"
        assert item["category_name"] == "工学"
        assert item["discipline_name"] == "计算机科学"
        assert item["existing_category_id"] is None
        assert result["summary"]["new"] == 1

    @pytest.mark.asyncio
    @patch(DISC_REPO)
    async def test_existing_category_new_discipline(
        self, mock_repo, service
    ):
        """大分类存在但小分类不存在，status 为 update。"""
        content = _create_workbook([["工学", "量子计算"]])
        cat = _make_category()
        mock_repo.get_category_by_name = AsyncMock(
            return_value=cat
        )
        mock_repo.get_discipline_by_name = AsyncMock(
            return_value=None
        )

        result = await service.preview(content)

        assert len(result["items"]) == 1
        assert result["items"][0]["status"] == "update"
        assert (
            result["items"][0]["existing_category_id"]
            == "cat-1"
        )
        assert result["summary"]["update"] == 1

    @pytest.mark.asyncio
    @patch(DISC_REPO)
    async def test_existing_both_unchanged(
        self, mock_repo, service
    ):
        """大分类和小分类都存在，status 为 unchanged。"""
        content = _create_workbook([["工学", "计算机科学"]])
        cat = _make_category()
        disc = _make_discipline()
        mock_repo.get_category_by_name = AsyncMock(
            return_value=cat
        )
        mock_repo.get_discipline_by_name = AsyncMock(
            return_value=disc
        )

        result = await service.preview(content)

        assert len(result["items"]) == 1
        assert result["items"][0]["status"] == "unchanged"
        assert (
            result["items"][0]["existing_discipline_id"]
            == "disc-1"
        )
        assert result["summary"]["unchanged"] == 1

    @pytest.mark.asyncio
    @patch(DISC_REPO)
    async def test_empty_row_reports_error(
        self, mock_repo, service
    ):
        """大分类或小分类为空的行报错。"""
        content = _create_workbook([
            ["工学", ""],
            ["", "计算机科学"],
        ])

        result = await service.preview(content)

        assert len(result["items"]) == 0
        assert len(result["errors"]) == 2
        assert result["summary"]["error"] == 2

    @pytest.mark.asyncio
    @patch(DISC_REPO)
    async def test_mixed_statuses(
        self, mock_repo, service
    ):
        """多行混合状态：new + update + unchanged + error。"""
        content = _create_workbook([
            ["工学", "计算机科学"],
            ["工学", "量子计算"],
            ["文学", "古典文学"],
            ["", "空大分类"],
        ])

        cat = _make_category()
        disc = _make_discipline()

        async def mock_get_cat(session, name):
            """按名称返回分类。"""
            if name == "工学":
                return cat
            return None

        async def mock_get_disc(session, cat_id, name):
            """按名称返回学科。"""
            if name == "计算机科学":
                return disc
            return None

        mock_repo.get_category_by_name = mock_get_cat
        mock_repo.get_discipline_by_name = mock_get_disc

        result = await service.preview(content)

        assert result["summary"]["unchanged"] == 1
        assert result["summary"]["update"] == 1
        assert result["summary"]["new"] == 1
        assert result["summary"]["error"] == 1

    @pytest.mark.asyncio
    @patch(DISC_REPO)
    async def test_short_row_reports_error(
        self, mock_repo, service
    ):
        """不足两列的行报错。"""
        content = _create_workbook([["只有一列"]])

        result = await service.preview(content)

        assert len(result["errors"]) == 1
        assert "不能为空" in result["errors"][0]["error"]
