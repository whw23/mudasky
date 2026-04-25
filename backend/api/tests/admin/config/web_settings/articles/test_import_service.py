"""ArticleImportService 单元测试 — generate_template / _parse_workbook。

测试文章批量导入的模板生成和 Excel 解析逻辑。
"""

import io
import zipfile
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import Workbook, load_workbook

from api.admin.config.web_settings.articles.import_service import (
    ArticleImportService,
)

from .conftest_import import (
    CONTENT_REPO,
    create_valid_workbook,
    make_article,
)


@pytest.fixture
def service(mock_session) -> ArticleImportService:
    """构建 ArticleImportService 实例。"""
    return ArticleImportService(mock_session)


# ---- generate_template ----


def test_generate_template_returns_valid_zip(service):
    """生成导入模板返回有效的 ZIP 文件。"""
    result = service.generate_template()

    assert isinstance(result, bytes)
    assert len(result) > 0

    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        names = zf.namelist()
        assert any(n.endswith(".xlsx") for n in names)
        assert any(n.startswith("content/") for n in names)


def test_generate_template_has_required_headers(service):
    """生成的模板包含必填字段列头。"""
    result = service.generate_template()

    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        xlsx = [n for n in zf.namelist() if n.endswith(".xlsx")][0]
        wb = load_workbook(io.BytesIO(zf.read(xlsx)))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "标题" in headers
        assert "内容类型" in headers
        assert "正文文件名" in headers


def test_generate_template_has_sample_rows(service):
    """生成的模板包含示例数据行。"""
    result = service.generate_template()

    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        xlsx = [n for n in zf.namelist() if n.endswith(".xlsx")][0]
        wb = load_workbook(io.BytesIO(zf.read(xlsx)))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 2
        assert rows[0][0] == "德国留学完整指南"


# ---- _parse_workbook ----


@pytest.mark.asyncio
@patch(CONTENT_REPO)
async def test_parse_workbook_new_article(mock_repo, service):
    """解析有效 workbook，新文章返回 status=new。"""
    wb = create_valid_workbook()
    mock_repo.get_article_by_title = AsyncMock(return_value=None)

    items, errors = await service._parse_workbook(wb, "cat-1")

    assert len(items) == 1
    assert items[0]["title"] == "德国留学指南"
    assert items[0]["status"] == "new"
    assert items[0]["data"]["status"] == "published"
    assert items[0]["data"]["is_pinned"] is True
    assert len(errors) == 0


@pytest.mark.asyncio
@patch(CONTENT_REPO)
async def test_parse_workbook_existing_update(mock_repo, service):
    """已存在且有变化的文章返回 status=update。"""
    wb = create_valid_workbook()
    existing = make_article(
        status="draft", is_pinned=False, excerpt="旧摘要"
    )
    mock_repo.get_article_by_title = AsyncMock(
        return_value=existing
    )

    items, errors = await service._parse_workbook(wb, "cat-1")

    assert len(items) == 1
    assert items[0]["status"] == "update"
    assert len(items[0]["changed_fields"]) > 0


@pytest.mark.asyncio
@patch(CONTENT_REPO)
async def test_parse_workbook_missing_content_filename(
    mock_repo, service
):
    """正文文件名为空时报错。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "文章"
    ws.append([
        "标题", "内容类型", "正文文件名", "摘要",
        "封面图文件名", "状态", "是否置顶",
    ])
    ws.append([
        "德国留学指南", "html", None, "摘要",
        None, "draft", "否",
    ])

    existing = make_article()
    mock_repo.get_article_by_title = AsyncMock(
        return_value=existing
    )

    items, errors = await service._parse_workbook(wb, "cat-1")

    assert len(errors) == 1
    assert "正文文件名" in errors[0]["error"]


@pytest.mark.asyncio
@patch(CONTENT_REPO)
async def test_parse_workbook_empty_row_skipped(mock_repo, service):
    """空行被跳过。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "文章"
    ws.append(["标题", "内容类型", "正文文件名"])
    ws.append([None, None, None])

    items, errors = await service._parse_workbook(wb, "cat-1")

    assert len(items) == 0
    assert len(errors) == 0


@pytest.mark.asyncio
@patch(CONTENT_REPO)
async def test_parse_workbook_invalid_content_type(
    mock_repo, service
):
    """无效内容类型报错。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "文章"
    ws.append([
        "标题", "内容类型", "正文文件名", "摘要",
        "封面图文件名", "状态", "是否置顶",
    ])
    ws.append([
        "标题", "video", "file.mp4", "摘要",
        None, "draft", "否",
    ])

    mock_repo.get_article_by_title = AsyncMock(return_value=None)

    items, errors = await service._parse_workbook(wb, "cat-1")

    assert len(items) == 0
    assert len(errors) == 1
    assert "内容类型" in errors[0]["error"]


@pytest.mark.asyncio
@patch(CONTENT_REPO)
async def test_parse_workbook_default_content_type(
    mock_repo, service
):
    """内容类型为空时默认 html。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "文章"
    ws.append([
        "标题", "内容类型", "正文文件名", "摘要",
        "封面图文件名", "状态", "是否置顶",
    ])
    ws.append([
        "默认类型文章", None, "file.html", "摘要",
        None, "draft", "否",
    ])

    mock_repo.get_article_by_title = AsyncMock(return_value=None)

    items, errors = await service._parse_workbook(wb, "cat-1")

    assert len(items) == 1
    assert items[0]["data"]["content_type"] == "html"


# ---- _guess_mime_type ----


def test_guess_mime_type_jpg(service):
    """jpg 返回 image/jpeg。"""
    assert service._guess_mime_type("photo.jpg") == "image/jpeg"


def test_guess_mime_type_png(service):
    """png 返回 image/png。"""
    assert service._guess_mime_type("icon.png") == "image/png"


def test_guess_mime_type_unknown(service):
    """未知扩展名返回 image/jpeg。"""
    assert service._guess_mime_type("file.bmp") == "image/jpeg"


def test_guess_mime_type_svg(service):
    """svg 返回 image/svg+xml。"""
    assert service._guess_mime_type("logo.svg") == "image/svg+xml"
