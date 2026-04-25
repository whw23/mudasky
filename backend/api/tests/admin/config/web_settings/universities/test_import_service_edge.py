"""ImportService 补充测试 - 边界情况。

覆盖 _parse_workbook、_find_unknown_disciplines、preview ZIP 无 Excel。
"""

import io
import zipfile
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook

from api.admin.config.web_settings.universities.import_service import (
    ImportService,
)
from app.db.discipline.models import Discipline, DisciplineCategory

DISC_REPO = (
    "api.admin.config.web_settings.universities"
    ".import_service.disc_repo"
)
UNI_REPO = (
    "api.admin.config.web_settings.universities"
    ".import_service.uni_repo"
)


def _make_category(
    category_id: str = "cat-1", name: str = "工学"
) -> MagicMock:
    """创建模拟 DisciplineCategory。"""
    cat = MagicMock(spec=DisciplineCategory)
    cat.id = category_id
    cat.name = name
    return cat


def _make_upload_file(
    content: bytes, filename: str = "test.xlsx"
) -> MagicMock:
    """创建模拟 UploadFile。"""
    file = MagicMock()
    file.filename = filename
    file.read = AsyncMock(return_value=content)
    return file


@pytest.fixture
def service(mock_session) -> ImportService:
    """构建 ImportService 实例。"""
    return ImportService(mock_session)


class TestEdgeCases:
    """边界情况测试。"""

    @pytest.mark.asyncio
    async def test_missing_basic_info_sheet(self, service):
        """缺少基本信息 sheet 时返回错误。"""
        wb = Workbook()
        wb.active.title = "其他"
        items, errors = await service._parse_workbook(wb, {})
        assert len(errors) == 1
        assert "基本信息" in errors[0]["error"]

    @pytest.mark.asyncio
    @patch(UNI_REPO)
    async def test_empty_program_rows_skipped(
        self, mock_u, service
    ):
        """专业列表中空行被跳过。"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "基本信息"
        ws1.append(["名称"] + [""] * 15)
        ws1.append(
            ["T", None, "中国", None, "北京"] + [None] * 11
        )
        ws2 = wb.create_sheet("专业列表")
        ws2.append(["院校名称", "专业名称", "大分类", "小分类"])
        ws2.append([None, "CS", "工学", "CS"])
        ws2.append(["T", None, "工学", "CS"])
        mock_u.get_university_by_name = AsyncMock(
            return_value=None
        )
        items, _ = await service._parse_workbook(wb, {})
        assert items[0]["programs"] == []

    @pytest.mark.asyncio
    @patch(DISC_REPO)
    async def test_invalid_disc_format(
        self, mock_repo, service
    ):
        """无斜杠路径标记为未知。"""
        result = await service._find_unknown_disciplines(
            {"无斜杠"}
        )
        assert "无斜杠" in result

    @pytest.mark.asyncio
    @patch(DISC_REPO)
    async def test_unknown_disc_in_known_cat(
        self, mock_repo, service
    ):
        """已知分类下的未知学科被标记。"""
        mock_repo.get_category_by_name = AsyncMock(
            return_value=_make_category()
        )
        mock_repo.get_discipline_by_name = AsyncMock(
            return_value=None
        )
        result = await service._find_unknown_disciplines(
            {"工学/量子计算"}
        )
        assert "工学/量子计算" in result

    @pytest.mark.asyncio
    async def test_zip_without_excel_raises(self, service):
        """ZIP 中无 xlsx 文件时抛出异常。"""
        from app.core.exceptions import BadRequestException

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w") as zf:
            zf.writestr("images/logo.jpg", b"data")
        file = _make_upload_file(
            zip_buf.getvalue(), "batch.zip"
        )
        with pytest.raises(BadRequestException):
            await service.preview(file)
