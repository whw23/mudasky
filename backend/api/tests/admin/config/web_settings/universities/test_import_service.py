"""ImportService 单元测试。

测试院校批量导入的业务逻辑。
"""

import io
import zipfile
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook, load_workbook

from api.admin.config.web_settings.universities.import_service import ImportService
from app.db.discipline.models import Discipline, DisciplineCategory
from app.db.university.models import University

DISC_REPO = "api.admin.config.web_settings.universities.import_service.disc_repo"
UNI_REPO = "api.admin.config.web_settings.universities.import_service.uni_repo"
PROG_REPO = "api.admin.config.web_settings.universities.import_service.prog_repo"
IMAGE_REPO = "api.admin.config.web_settings.universities.import_service.image_repo"


def _make_category(category_id: str = "cat-1", name: str = "工学") -> MagicMock:
    """创建模拟 DisciplineCategory 对象。"""
    cat = MagicMock(spec=DisciplineCategory)
    cat.id = category_id
    cat.name = name
    return cat


def _make_discipline(
    discipline_id: str = "disc-1", name: str = "计算机科学"
) -> MagicMock:
    """创建模拟 Discipline 对象。"""
    disc = MagicMock(spec=Discipline)
    disc.id = discipline_id
    disc.name = name
    return disc


def _make_university(university_id: str = "uni-1") -> MagicMock:
    """创建模拟 University 对象。"""
    uni = MagicMock(spec=University)
    uni.id = university_id
    uni.name = "哈佛大学"
    uni.name_en = "Harvard University"
    uni.country = "美国"
    uni.province = "马萨诸塞州"
    uni.city = "剑桥"
    uni.website = "https://harvard.edu"
    uni.description = "世界顶尖学府"
    uni.admission_requirements = None
    uni.scholarship_info = None
    uni.latitude = None
    uni.longitude = None
    uni.is_featured = False
    uni.sort_order = 0
    uni.qs_rankings = None
    uni.logo_image_id = None
    return uni


def _make_upload_file(content: bytes, filename: str = "test.xlsx") -> MagicMock:
    """创建模拟 UploadFile 对象。"""
    file = MagicMock()
    file.filename = filename
    file.read = AsyncMock(return_value=content)
    return file


def _create_valid_workbook() -> Workbook:
    """创建符合实际格式的横向 workbook。"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "基本信息"
    ws1.append([
        "名称", "英文名", "国家", "省份", "城市", "网站", "描述",
        "录取要求", "奖学金信息", "纬度", "经度", "是否精选", "排序",
        "Logo文件名", "展示图文件名", "QS排名",
    ])
    ws1.append([
        "哈佛大学", "Harvard University", "美国", "马萨诸塞州", "剑桥",
        "https://harvard.edu", "世界顶尖学府", "GPA 3.8+", "多种奖学金",
        42.377, -71.1167, "是", 0, None, None,
        '[{"year": 2026, "ranking": 4}]',
    ])

    ws2 = wb.create_sheet("专业列表")
    ws2.append(["院校名称", "专业名称", "大分类", "小分类"])
    ws2.append(["哈佛大学", "计算机科学", "工学", "计算机科学"])
    ws2.append(["哈佛大学", "金融学", "商学", "金融学"])

    return wb


def _workbook_to_bytes(wb: Workbook) -> bytes:
    """将 workbook 转换为字节。"""
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


@pytest.fixture
def service(mock_session) -> ImportService:
    """构建 ImportService 实例。"""
    return ImportService(mock_session)


# ---- generate_template ----


def test_generate_template(service):
    """生成导入模板返回有效的 ZIP 文件。"""
    result = service.generate_template()

    assert isinstance(result, bytes)
    assert len(result) > 0

    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        names = zf.namelist()
        assert any(n.endswith(".xlsx") for n in names)

        excel_data = zf.read([n for n in names if n.endswith(".xlsx")][0])
        wb = load_workbook(io.BytesIO(excel_data))
        assert "基本信息" in wb.sheetnames
        assert "专业列表" in wb.sheetnames


def test_generate_template_has_required_fields(service):
    """生成的模板包含必填字段列头。"""
    result = service.generate_template()

    with zipfile.ZipFile(io.BytesIO(result)) as zf:
        excel_data = zf.read([n for n in zf.namelist() if n.endswith(".xlsx")][0])
        wb = load_workbook(io.BytesIO(excel_data))
        ws1 = wb["基本信息"]

        headers = [cell.value for cell in ws1[1]]
        assert "名称" in headers
        assert "国家" in headers
        assert "城市" in headers


# ---- _parse_workbook ----


@pytest.mark.asyncio
@patch(UNI_REPO)
async def test_parse_workbook_success(mock_uni_repo, service):
    """解析有效 workbook 返回 items 列表。"""
    wb = _create_valid_workbook()
    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)

    items, errors = await service._parse_workbook(wb, {})

    assert len(items) == 1
    assert items[0]["name"] == "哈佛大学"
    assert items[0]["status"] == "new"
    assert items[0]["data"]["country"] == "美国"
    assert items[0]["data"]["city"] == "剑桥"
    assert len(items[0]["programs"]) == 2
    assert len(errors) == 0


@pytest.mark.asyncio
@patch(UNI_REPO)
async def test_parse_workbook_missing_required_name(mock_uni_repo, service):
    """名称为空的行被跳过（不报错）。"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "基本信息"
    ws1.append(["名称", "英文名", "国家", "省份", "城市"] + [""] * 11)
    ws1.append([None, None, "美国", None, "剑桥"] + [None] * 11)

    items, errors = await service._parse_workbook(wb, {})

    assert len(items) == 0


@pytest.mark.asyncio
@patch(UNI_REPO)
async def test_parse_workbook_missing_required_country(mock_uni_repo, service):
    """缺少国家的行报错。"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "基本信息"
    ws1.append(["名称", "英文名", "国家", "省份", "城市"] + [""] * 11)
    ws1.append(["哈佛大学", None, None, None, "剑桥"] + [None] * 11)

    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)

    items, errors = await service._parse_workbook(wb, {})

    assert len(items) == 0
    assert len(errors) == 1
    assert "国家" in errors[0]["error"]


@pytest.mark.asyncio
@patch(UNI_REPO)
async def test_parse_workbook_missing_required_city(mock_uni_repo, service):
    """缺少城市的行报错。"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "基本信息"
    ws1.append(["名称", "英文名", "国家", "省份", "城市"] + [""] * 11)
    ws1.append(["哈佛大学", None, "美国", None, None] + [None] * 11)

    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)

    items, errors = await service._parse_workbook(wb, {})

    assert len(items) == 0
    assert len(errors) == 1
    assert "城市" in errors[0]["error"]


@pytest.mark.asyncio
@patch(UNI_REPO)
async def test_parse_workbook_without_optional_sheets(mock_uni_repo, service):
    """没有专业列表 sheet 也能解析。"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "基本信息"
    ws1.append(["名称", "英文名", "国家", "省份", "城市"] + [""] * 11)
    ws1.append(["哈佛大学", None, "美国", None, "剑桥"] + [None] * 11)

    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)

    items, errors = await service._parse_workbook(wb, {})

    assert len(items) == 1
    assert items[0]["programs"] == []


# ---- preview ----


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(DISC_REPO)
async def test_preview_single_file(mock_disc_repo, mock_uni_repo, service):
    """预览单个 xlsx 返回 items/errors/summary。"""
    wb = _create_valid_workbook()
    content = _workbook_to_bytes(wb)
    file = _make_upload_file(content, "test.xlsx")

    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)
    cat = _make_category()
    disc1 = _make_discipline("disc-1", "计算机科学")
    disc2 = _make_discipline("disc-2", "金融学")
    mock_disc_repo.get_category_by_name = AsyncMock(return_value=cat)
    mock_disc_repo.get_discipline_by_name = AsyncMock(
        side_effect=[disc1, disc2]
    )

    result = await service.preview(file)

    assert len(result["items"]) == 1
    assert result["items"][0]["name"] == "哈佛大学"
    assert result["summary"]["new"] == 1
    assert len(result["errors"]) == 0


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(DISC_REPO)
async def test_preview_with_unknown_disciplines(
    mock_disc_repo, mock_uni_repo, service
):
    """预览时检测到未知学科。"""
    wb = _create_valid_workbook()
    content = _workbook_to_bytes(wb)
    file = _make_upload_file(content, "test.xlsx")

    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)

    async def mock_get_cat(session, name):
        if name == "工学":
            return _make_category("cat-1", "工学")
        return None

    async def mock_get_disc(session, cat_id, name):
        if cat_id == "cat-1" and name == "计算机科学":
            return _make_discipline("disc-1", "计算机科学")
        return None

    mock_disc_repo.get_category_by_name = mock_get_cat
    mock_disc_repo.get_discipline_by_name = mock_get_disc

    result = await service.preview(file)

    assert len(result["unknown_disciplines"]) >= 1
    assert "商学/金融学" in result["unknown_disciplines"]


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(DISC_REPO)
async def test_preview_with_unknown_category(
    mock_disc_repo, mock_uni_repo, service
):
    """预览时检测到未知学科分类。"""
    wb = _create_valid_workbook()
    content = _workbook_to_bytes(wb)
    file = _make_upload_file(content, "test.xlsx")

    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)
    mock_disc_repo.get_category_by_name = AsyncMock(return_value=None)

    result = await service.preview(file)

    assert len(result["unknown_disciplines"]) == 2
    assert "工学/计算机科学" in result["unknown_disciplines"]
    assert "商学/金融学" in result["unknown_disciplines"]


@pytest.mark.asyncio
@patch(UNI_REPO)
async def test_preview_invalid_file(mock_uni_repo, service):
    """缺少必填字段的行进入 errors。"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "基本信息"
    ws1.append(["名称", "英文名", "国家", "省份", "城市"] + [""] * 11)
    ws1.append(["哈佛大学", None, None, None, None] + [None] * 11)

    content = _workbook_to_bytes(wb)
    file = _make_upload_file(content, "test.xlsx")
    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)

    result = await service.preview(file)

    assert len(result["items"]) == 0
    assert len(result["errors"]) == 1


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(DISC_REPO)
async def test_preview_multiple_files_in_zip(
    mock_disc_repo, mock_uni_repo, service
):
    """ZIP 中只解析第一个 xlsx。"""
    wb = _create_valid_workbook()

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zf:
        zf.writestr("universities.xlsx", _workbook_to_bytes(wb))
        zf.writestr("images/.gitkeep", b"")

    file = _make_upload_file(zip_buffer.getvalue(), "batch.zip")

    mock_uni_repo.get_university_by_name = AsyncMock(return_value=None)
    cat = _make_category()
    disc1 = _make_discipline("disc-1", "计算机科学")
    disc2 = _make_discipline("disc-2", "金融学")
    mock_disc_repo.get_category_by_name = AsyncMock(return_value=cat)
    mock_disc_repo.get_discipline_by_name = AsyncMock(
        side_effect=[disc1, disc2]
    )

    result = await service.preview(file)

    assert len(result["items"]) == 1
    assert result["items"][0]["name"] == "哈佛大学"


# ---- confirm ----


@pytest.mark.asyncio
@patch(PROG_REPO)
@patch(UNI_REPO)
@patch(DISC_REPO)
async def test_confirm_import_success(
    mock_disc_repo, mock_uni_repo, mock_prog_repo, service
):
    """确认导入创建新院校。"""
    items = [{
        "name": "哈佛大学",
        "status": "new",
        "data": {"name": "哈佛大学", "country": "美国", "city": "剑桥"},
        "programs": [],
        "logo_filename": None,
        "image_filenames": [],
    }]

    uni = _make_university()
    mock_uni_repo.create_university = AsyncMock(return_value=uni)

    file = _make_upload_file(b"fake", "test.xlsx")
    result = await service.confirm(file, items, [])

    assert result["imported"] == 1
    assert result["skipped"] == 0
    mock_uni_repo.create_university.assert_awaited_once()


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(DISC_REPO)
async def test_confirm_with_discipline_mappings(
    mock_disc_repo, mock_uni_repo, service
):
    """确认导入时创建新学科分类。"""
    items = [{
        "name": "MIT",
        "status": "new",
        "data": {"name": "MIT", "country": "美国", "city": "剑桥"},
        "programs": [],
        "logo_filename": None,
        "image_filenames": [],
    }]

    mappings = [{"action": "create", "name": "理学/物理学"}]

    cat = _make_category("cat-2", "理学")
    disc = _make_discipline("disc-3", "物理学")
    uni = _make_university("uni-2")

    mock_disc_repo.get_category_by_name = AsyncMock(return_value=None)
    mock_disc_repo.create_category = AsyncMock(return_value=cat)
    mock_disc_repo.get_discipline_by_name = AsyncMock(return_value=None)
    mock_disc_repo.create_discipline = AsyncMock(return_value=disc)
    mock_uni_repo.create_university = AsyncMock(return_value=uni)

    file = _make_upload_file(b"fake", "test.xlsx")
    result = await service.confirm(file, items, mappings)

    assert result["imported"] == 1
    mock_disc_repo.create_category.assert_awaited_once()
    mock_disc_repo.create_discipline.assert_awaited_once()


@pytest.mark.asyncio
@patch(UNI_REPO)
@patch(DISC_REPO)
async def test_confirm_skip_failed_rows(
    mock_disc_repo, mock_uni_repo, service
):
    """导入时跳过失败的行。"""
    items = [
        {
            "name": "哈佛大学", "status": "new",
            "data": {"name": "哈佛大学", "country": "美国", "city": "剑桥"},
            "programs": [], "logo_filename": None, "image_filenames": [],
        },
        {
            "name": "MIT", "status": "new",
            "data": {"name": "MIT", "country": "美国", "city": "剑桥"},
            "programs": [], "logo_filename": None, "image_filenames": [],
        },
    ]

    uni = _make_university()
    mock_uni_repo.create_university = AsyncMock(
        side_effect=[uni, Exception("DB error")]
    )

    file = _make_upload_file(b"fake", "test.xlsx")
    result = await service.confirm(file, items, [])

    assert result["imported"] == 1
    assert result["skipped"] == 1
