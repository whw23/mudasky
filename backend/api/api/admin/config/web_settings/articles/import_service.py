"""文章批量导入服务。

支持 ZIP（包含 Excel + content/）和纯 Excel 两种格式。
Excel 格式：单个 sheet，横向表
  标题 | 内容类型 | 正文文件名 | 摘要 | 封面图文件名 | 状态 | 是否置顶

导入策略：merge（标题+分类为主键，非空字段覆盖），slug 从标题自动生成。
特点：
- category_id 通过 query param 传入，不在 Excel 中
- author_id 使用当前登录用户 ID（从请求头 X-User-Id 获取）
- content_type 为 html 时，正文从 HTML 文件读取；为 file 时，正文为 PDF
- cover_image 存为 URL 字符串（/api/public/images/detail?id={image_id}）
"""

import io

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import BadRequestException
from app.db.content import repository
from app.db.content.models import Article
from app.db.image import repository as image_repo
from app.utils.slug import generate_unique_slug
from app.utils.excel_io import (
    create_placeholder_image,
    create_zip,
    extract_zip,
    load_workbook_from_bytes,
    workbook_to_bytes,
    write_sheet_header,
)


class ArticleImportService:
    """文章批量导入服务。"""

    def __init__(self, session: AsyncSession) -> None:
        """初始化服务。"""
        self.session = session

    async def preview(
        self, content: bytes, category_id: str, is_zip: bool = False
    ) -> dict:
        """解析上传文件，返回预览结果。

        返回格式：
        {
          "items": [{"row": 2, "title": "...", "status": "new|update|unchanged", "changed_fields": [...], "data": {...}}],
          "errors": [{"row": 3, "error": "..."}],
          "summary": {"new": 1, "update": 2, "unchanged": 0, "error": 1},
          "available_files": ["content/study-in-germany.html", "content/cover1.jpg"]
        }
        """
        # 解析 ZIP 或 Excel
        if is_zip:
            files = extract_zip(content)
            excel_data = None
            for name, data in files.items():
                if name.endswith(".xlsx") and not name.startswith("__"):
                    excel_data = data
                    break
            if not excel_data:
                raise BadRequestException(
                    message="ZIP 中未找到 Excel 文件",
                    code="NO_EXCEL_IN_ZIP",
                )
            wb = load_workbook_from_bytes(excel_data)
            available_files = [
                name for name in files.keys() if name.startswith("content/")
            ]
        else:
            wb = load_workbook_from_bytes(content)
            available_files = []

        # 解析 Excel
        items, errors = await self._parse_workbook(wb, category_id)

        summary = {
            "new": sum(1 for x in items if x["status"] == "new"),
            "update": sum(1 for x in items if x["status"] == "update"),
            "unchanged": sum(1 for x in items if x["status"] == "unchanged"),
            "error": len(errors),
        }

        return {
            "items": items,
            "errors": errors,
            "summary": summary,
            "available_files": available_files,
        }

    async def confirm(
        self,
        items: list[dict],
        author_id: str,
        file_content: bytes,
        is_zip: bool = False,
    ) -> dict:
        """执行批量导入。

        Args:
            items: preview 返回的 items
            author_id: 当前登录用户 ID（从 X-User-Id 获取）
            file_content: 原始上传文件内容（重新解析获取文件）
            is_zip: 是否为 ZIP 格式

        Returns:
            {"imported": 1, "updated": 2, "skipped": 3}
        """
        # 重新解析文件获取内容文件和图片
        file_map = {}
        if is_zip:
            files = extract_zip(file_content)
            file_map = {
                name: files[name]
                for name in files.keys()
                if name.startswith("content/")
            }

        imported = 0
        updated = 0
        skipped = 0

        for item in items:
            try:
                if item["status"] == "new":
                    await self._create_article(item, author_id, file_map)
                    imported += 1
                elif item["status"] == "update":
                    await self._update_article(item, file_map)
                    updated += 1
                else:  # unchanged
                    skipped += 1
            except Exception:
                skipped += 1

        return {"imported": imported, "updated": updated, "skipped": skipped}

    def generate_template(self) -> bytes:
        """生成 ZIP 导入模板（包含 Excel + 空 content/ 目录）。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "文章"
        headers = [
            "标题",
            "内容类型",
            "正文文件名",
            "摘要",
            "封面图文件名",
            "状态",
            "是否置顶",
        ]
        write_sheet_header(ws, headers)
        # 示例行
        ws.append(
            [
                "德国留学完整指南",
                "html",
                "study-in-germany.html",
                "本文介绍德国留学的各个方面...",
                "cover1.jpg",
                "published",
                "是",
            ]
        )
        ws.append(
            [
                "签证申请注意事项",
                "file",
                "visa-checklist.pdf",
                "签证申请流程和材料清单",
                "cover2.jpg",
                "draft",
                "否",
            ]
        )

        excel_bytes = workbook_to_bytes(wb)

        # 创建 ZIP
        files = {
            "articles.xlsx": excel_bytes,
            "content/study-in-germany.html": b"<h1>Study in Germany</h1><p>Placeholder content.</p>",
            "content/cover1.jpg": create_placeholder_image(800, 450, "Cover 1"),
            "content/cover2.jpg": create_placeholder_image(800, 450, "Cover 2"),
        }
        return create_zip(files)

    async def _parse_workbook(
        self, wb: Workbook, category_id: str
    ) -> tuple[list[dict], list[dict]]:
        """解析 workbook，返回 (items, errors)。"""
        items = []
        errors = []

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for idx, row in enumerate(rows, start=2):
            if not row[0]:
                continue  # 跳过空行

            try:
                item = await self._parse_row(idx, row, category_id)
                items.append(item)
            except ValueError as e:
                errors.append({"row": idx, "error": str(e)})

        return items, errors

    async def _parse_row(
        self, row_num: int, row: tuple, category_id: str
    ) -> dict:
        """解析单行为 item。

        返回格式：
        {
          "row": 2,
          "title": "...",
          "status": "new|update|unchanged",
          "changed_fields": ["title", "content"],
          "data": {"title": "...", ...},
          "content_filename": "study-guide.html",
          "cover_image_filename": "cover1.jpg",
        }
        """
        title = str(row[0]).strip() if row[0] else ""
        if not title:
            raise ValueError("标题不能为空")

        content_type = str(row[1]).strip().lower() if row[1] else "html"
        if content_type not in ["html", "file"]:
            raise ValueError(f"行 {row_num}: 内容类型必须为 html 或 file")

        content_filename = str(row[2]).strip() if row[2] else None
        excerpt = str(row[3]).strip() if row[3] else ""
        cover_image_filename = str(row[4]).strip() if row[4] else None

        status = "draft"
        if row[5]:
            val = str(row[5]).strip().lower()
            if val in ["published", "已发布"]:
                status = "published"

        is_pinned = False
        if row[6]:
            val = str(row[6]).strip().lower()
            is_pinned = val in ["是", "true", "1", "yes"]

        if not content_filename:
            raise ValueError(f"行 {row_num}: 正文文件名不能为空")

        # 按标题匹配已有文章
        existing = await repository.get_article_by_title(
            self.session, title, category_id,
        )
        data = {
            "title": title,
            "content_type": content_type,
            "excerpt": excerpt,
            "category_id": category_id,
            "status": status,
            "is_pinned": is_pinned,
        }

        if not existing:
            return {
                "row": row_num,
                "title": title,
                "status": "new",
                "changed_fields": [],
                "data": data,
                "content_filename": content_filename,
                "cover_image_filename": cover_image_filename,
            }

        changed_fields = []
        for key, new_val in data.items():
            if key in ["category_id"]:
                continue
            if new_val is None or new_val == "":
                continue
            old_val = getattr(existing, key, None)
            if new_val != old_val:
                changed_fields.append(key)

        if content_filename:
            changed_fields.append("content")
        if cover_image_filename:
            changed_fields.append("cover_image")

        status_result = "update" if changed_fields else "unchanged"

        return {
            "row": row_num,
            "title": title,
            "status": status_result,
            "changed_fields": changed_fields,
            "data": data,
            "content_filename": content_filename,
            "cover_image_filename": cover_image_filename,
        }

    async def _create_article(
        self, item: dict, author_id: str, file_map: dict[str, bytes]
    ) -> Article:
        """创建新文章。"""
        data = item["data"]

        # 处理正文内容
        content_filename = item.get("content_filename")
        content_text = ""
        file_id = None

        if data["content_type"] == "html":
            # 读取 HTML 文件
            content_path = f"content/{content_filename}"
            if content_path in file_map:
                content_text = file_map[content_path].decode("utf-8")
            else:
                raise ValueError(f"未找到 HTML 文件: {content_filename}")
        else:  # file (PDF)
            # 上传 PDF
            content_path = f"content/{content_filename}"
            if content_path in file_map:
                pdf_data = file_map[content_path]
                image = await image_repo.create_image(
                    self.session,
                    pdf_data,
                    content_filename,
                    "application/pdf",
                )
                file_id = image.id
                content_text = ""  # PDF 不存文本
            else:
                raise ValueError(f"未找到 PDF 文件: {content_filename}")

        # 处理封面图
        cover_image_url = None
        cover_image_filename = item.get("cover_image_filename")
        if cover_image_filename:
            cover_path = f"content/{cover_image_filename}"
            if cover_path in file_map:
                cover_data = file_map[cover_path]
                mime_type = self._guess_mime_type(cover_image_filename)
                image = await image_repo.create_image(
                    self.session,
                    cover_data,
                    cover_image_filename,
                    mime_type,
                )
                cover_image_url = f"/api/public/images/detail?id={image.id}"

        slug = await generate_unique_slug(
            self.session, data["title"], Article,
        )
        article = Article(
            title=data["title"],
            slug=slug,
            content_type=data["content_type"],
            content=content_text,
            file_id=file_id,
            excerpt=data.get("excerpt", ""),
            cover_image=cover_image_url,
            category_id=data["category_id"],
            author_id=author_id,
            status=data.get("status", "draft"),
            is_pinned=data.get("is_pinned", False),
        )
        return await repository.create_article(self.session, article)

    async def _update_article(
        self, item: dict, file_map: dict[str, bytes]
    ) -> Article:
        """更新文章（merge 策略）。"""
        data = item["data"]
        existing = await repository.get_article_by_title(
            self.session, data["title"], data["category_id"],
        )
        if not existing:
            raise ValueError(f"文章「{data['title']}」不存在")

        for key, new_val in data.items():
            if key in ["category_id"]:
                continue
            if new_val is None or new_val == "":
                continue  # 空值不覆盖
            setattr(existing, key, new_val)

        # 更新正文内容
        content_filename = item.get("content_filename")
        if content_filename:
            if existing.content_type == "html":
                content_path = f"content/{content_filename}"
                if content_path in file_map:
                    existing.content = file_map[content_path].decode("utf-8")
            else:  # file (PDF)
                content_path = f"content/{content_filename}"
                if content_path in file_map:
                    pdf_data = file_map[content_path]
                    image = await image_repo.create_image(
                        self.session,
                        pdf_data,
                        content_filename,
                        "application/pdf",
                    )
                    existing.file_id = image.id
                    existing.content = ""

        # 更新封面图
        cover_image_filename = item.get("cover_image_filename")
        if cover_image_filename:
            cover_path = f"content/{cover_image_filename}"
            if cover_path in file_map:
                cover_data = file_map[cover_path]
                mime_type = self._guess_mime_type(cover_image_filename)
                image = await image_repo.create_image(
                    self.session,
                    cover_data,
                    cover_image_filename,
                    mime_type,
                )
                existing.cover_image = f"/api/public/images/detail?id={image.id}"

        return await repository.update_article(self.session, existing)

    def _guess_mime_type(self, filename: str) -> str:
        """根据文件扩展名推测 MIME 类型。"""
        ext = filename.lower().split(".")[-1]
        mime_map = {
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "png": "image/png",
            "gif": "image/gif",
            "webp": "image/webp",
            "svg": "image/svg+xml",
        }
        return mime_map.get(ext, "image/jpeg")
