"""院校批量导入服务。

支持 ZIP（包含 Excel + images/）和纯 Excel 两种格式。
Excel 格式：
- Sheet1 "基本信息"：横向行，一行一个院校
- Sheet2 "专业列表"：纵向表，院校名称 + 专业名称 + 大分类 + 小分类

导入策略：merge（非空字段覆盖，空字段保留）。
"""

import io

from fastapi import UploadFile
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import BadRequestException
from app.db.discipline import repository as disc_repo
from app.db.discipline.models import Discipline, DisciplineCategory
from app.db.image import repository as image_repo
from app.db.university import repository as uni_repo
from app.db.university import program_repository as prog_repo
from app.db.university.image_models import UniversityImage
from app.db.university.models import University
from app.utils.excel_io import (
    create_placeholder_image,
    create_zip,
    extract_zip,
    load_workbook_from_bytes,
    workbook_to_bytes,
    write_sheet_header,
)


class ImportService:
    """院校批量导入服务。"""

    def __init__(self, session: AsyncSession) -> None:
        """初始化服务。"""
        self.session = session

    async def preview(self, file: UploadFile) -> dict:
        """解析上传文件，返回预览结果。

        返回格式：
        {
          "items": [{"row": 2, "name": "哈佛大学", "status": "new|update|unchanged", "changed_fields": [...], "data": {...}, "programs": [...]}],
          "errors": [{"row": 3, "error": "..."}],
          "summary": {"new": 1, "update": 2, "unchanged": 0, "error": 1},
          "unknown_disciplines": ["工学/量子计算"],
          "available_images": ["images/哈佛大学_logo.jpg"]
        }
        """
        content = await file.read()
        filename = file.filename or ""

        # 解析 ZIP 或 Excel
        if filename.endswith(".zip"):
            files = extract_zip(content)
            excel_data = None
            for name, data in files.items():
                if name.endswith(".xlsx") and not name.startswith("__"):
                    excel_data = data
                    break
            if not excel_data:
                raise BadRequestException(
                    message="ZIP 中未找到 Excel 文件",
                    code="NO_EXCEL_IN_ZIP",
                )
            wb = load_workbook_from_bytes(excel_data)
            available_images = [
                name for name in files.keys() if name.startswith("images/")
            ]
            image_data_map = {
                name: files[name] for name in available_images
            }  # 保存图片数据供 confirm 使用
        else:
            wb = load_workbook_from_bytes(content)
            available_images = []
            image_data_map = {}

        # 解析 Sheet1 基本信息 + Sheet2 专业列表
        items, errors = await self._parse_workbook(wb, image_data_map)
        all_disciplines = set()
        for item in items:
            for prog in item["programs"]:
                cat_name = prog.get("category_name")
                disc_name = prog.get("discipline_name")
                if cat_name and disc_name:
                    all_disciplines.add(f"{cat_name}/{disc_name}")

        unknown = await self._find_unknown_disciplines(all_disciplines)

        summary = {
            "new": sum(1 for x in items if x["status"] == "new"),
            "update": sum(1 for x in items if x["status"] == "update"),
            "unchanged": sum(1 for x in items if x["status"] == "unchanged"),
            "error": len(errors),
        }

        # 将 image_data_map 附加到 session state（实际上不方便，改为要求 confirm 重新上传文件）
        # 为简化实现，confirm 时要求重新上传同一文件

        return {
            "items": items,
            "errors": errors,
            "summary": summary,
            "unknown_disciplines": unknown,
            "available_images": available_images,
        }

    async def confirm(
        self,
        file: UploadFile,
        items: list[dict],
        discipline_actions: list[dict],
    ) -> dict:
        """执行批量导入。

        Args:
            file: 原始上传文件（重新解析获取图片）
            items: preview 返回的 items（只需 name + data + programs）
            discipline_actions: [{"name": "工学/量子计算", "action": "create|map", "target_id": "..."}]

        Returns:
            {"imported": 1, "updated": 2, "skipped": 3}
        """
        # 重新解析文件获取图片数据
        content = await file.read()
        filename = file.filename or ""
        image_data_map = {}
        if filename.endswith(".zip"):
            files = extract_zip(content)
            image_data_map = {
                name: files[name]
                for name in files.keys()
                if name.startswith("images/")
            }

        # 应用学科映射
        await self._apply_discipline_actions(discipline_actions)

        imported = 0
        updated = 0
        skipped = 0

        for item in items:
            try:
                if item["status"] == "new":
                    await self._create_university(item, image_data_map)
                    imported += 1
                elif item["status"] == "update":
                    await self._update_university(item, image_data_map)
                    updated += 1
                else:  # unchanged
                    skipped += 1
            except Exception:
                skipped += 1

        return {"imported": imported, "updated": updated, "skipped": skipped}

    def generate_template(self) -> bytes:
        """生成 ZIP 导入模板（包含 Excel + 空 images/ 目录）。"""
        wb = Workbook()

        # Sheet1: 基本信息（横向）
        ws1 = wb.active
        ws1.title = "基本信息"
        headers = [
            "名称",
            "英文名",
            "国家",
            "省份",
            "城市",
            "网站",
            "描述",
            "录取要求",
            "奖学金信息",
            "纬度",
            "经度",
            "是否精选",
            "排序",
            "Logo文件名",
            "展示图文件名",
            "QS排名",
        ]
        write_sheet_header(ws1, headers)
        # 示例行
        ws1.append(
            [
                "哈佛大学",
                "Harvard University",
                "美国",
                "马萨诸塞州",
                "剑桥",
                "https://harvard.edu",
                "世界顶尖学府",
                "GPA 3.8+, TOEFL 100+",
                "多种奖学金可申请",
                42.377,
                -71.1167,
                "是",
                0,
                "哈佛大学_logo.jpg",
                "哈佛大学_1.jpg,哈佛大学_2.jpg",
                "[{\"year\": 2026, \"ranking\": 4}]",
            ]
        )

        # Sheet2: 专业列表（纵向）
        ws2 = wb.create_sheet("专业列表")
        write_sheet_header(ws2, ["院校名称", "专业名称", "大分类", "小分类"])
        ws2.append(["哈佛大学", "计算机科学", "工学", "计算机科学"])
        ws2.append(["哈佛大学", "金融学", "商学", "金融学"])

        excel_bytes = workbook_to_bytes(wb)

        # 创建 ZIP
        files = {
            "universities.xlsx": excel_bytes,
            "images/哈佛大学_logo.jpg": create_placeholder_image(200, 200, "Logo", "#6366f1"),
            "images/哈佛大学_1.jpg": create_placeholder_image(800, 450, "Photo 1"),
            "images/哈佛大学_2.jpg": create_placeholder_image(800, 450, "Photo 2"),
        }
        return create_zip(files)

    async def _parse_workbook(
        self, wb: Workbook, image_data_map: dict[str, bytes]
    ) -> tuple[list[dict], list[dict]]:
        """解析 workbook，返回 (items, errors)。"""
        items = []
        errors = []

        if "基本信息" not in wb.sheetnames:
            errors.append({"row": 0, "error": "缺少 Sheet '基本信息'"})
            return items, errors

        ws1 = wb["基本信息"]
        rows = list(ws1.iter_rows(min_row=2, values_only=True))

        # 解析 Sheet2 专业列表
        programs_map = {}  # {院校名称: [{"category_name": ..., "discipline_name": ..., "program_name": ...}]}
        if "专业列表" in wb.sheetnames:
            ws2 = wb["专业列表"]
            for row in ws2.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                uni_name = str(row[0]).strip()
                prog_name = str(row[1]).strip() if row[1] else ""
                cat_name = str(row[2]).strip() if row[2] else ""
                disc_name = str(row[3]).strip() if row[3] else ""
                if not uni_name or not prog_name:
                    continue
                if uni_name not in programs_map:
                    programs_map[uni_name] = []
                programs_map[uni_name].append(
                    {
                        "program_name": prog_name,
                        "category_name": cat_name,
                        "discipline_name": disc_name,
                    }
                )

        # 解析每一行
        for idx, row in enumerate(rows, start=2):
            if not row[0]:
                continue  # 跳过空行

            try:
                item = await self._parse_row(
                    idx, row, programs_map, image_data_map
                )
                items.append(item)
            except ValueError as e:
                errors.append({"row": idx, "error": str(e)})

        return items, errors

    async def _parse_row(
        self,
        row_num: int,
        row: tuple,
        programs_map: dict,
        image_data_map: dict[str, bytes],
    ) -> dict:
        """解析单行为 item。

        返回格式：
        {
          "row": 2,
          "name": "哈佛大学",
          "status": "new|update|unchanged",
          "changed_fields": ["country", "city"],
          "data": {"name": "哈佛大学", "country": "美国", ...},
          "programs": [{"program_name": "...", "category_name": "...", "discipline_name": "..."}],
          "logo_filename": "哈佛大学_logo.jpg",
          "image_filenames": ["哈佛大学_1.jpg", "哈佛大学_2.jpg"],
        }
        """
        name = str(row[0]).strip() if row[0] else ""
        if not name:
            raise ValueError("名称不能为空")

        # 解析字段
        name_en = str(row[1]).strip() if row[1] else None
        country = str(row[2]).strip() if row[2] else None
        province = str(row[3]).strip() if row[3] else None
        city = str(row[4]).strip() if row[4] else None
        website = str(row[5]).strip() if row[5] else None
        description = str(row[6]).strip() if row[6] else None
        admission_requirements = str(row[7]).strip() if row[7] else None
        scholarship_info = str(row[8]).strip() if row[8] else None

        latitude = None
        if row[9]:
            try:
                latitude = float(row[9])
            except ValueError:
                pass

        longitude = None
        if row[10]:
            try:
                longitude = float(row[10])
            except ValueError:
                pass

        is_featured = False
        if row[11]:
            val = str(row[11]).strip().lower()
            is_featured = val in ["是", "true", "1", "yes"]

        sort_order = 0
        if row[12]:
            try:
                sort_order = int(row[12])
            except ValueError:
                pass

        logo_filename = str(row[13]).strip() if row[13] else None
        image_filenames_str = str(row[14]).strip() if row[14] else None
        image_filenames = []
        if image_filenames_str:
            image_filenames = [
                f.strip()
                for f in image_filenames_str.split(",")
                if f.strip()
            ]

        qs_rankings = None
        if row[15]:
            import json

            try:
                qs_rankings = json.loads(str(row[15]))
            except (json.JSONDecodeError, ValueError):
                pass

        # 必填字段校验
        if not country:
            raise ValueError(f"行 {row_num}: 国家不能为空")
        if not city:
            raise ValueError(f"行 {row_num}: 城市不能为空")

        # 查询数据库判断 new/update/unchanged
        existing = await uni_repo.get_university_by_name(self.session, name)
        data = {
            "name": name,
            "name_en": name_en,
            "country": country,
            "province": province,
            "city": city,
            "website": website,
            "description": description,
            "admission_requirements": admission_requirements,
            "scholarship_info": scholarship_info,
            "latitude": latitude,
            "longitude": longitude,
            "is_featured": is_featured,
            "sort_order": sort_order,
            "qs_rankings": qs_rankings,
        }

        programs = programs_map.get(name, [])

        if not existing:
            return {
                "row": row_num,
                "name": name,
                "status": "new",
                "changed_fields": [],
                "data": data,
                "programs": programs,
                "logo_filename": logo_filename,
                "image_filenames": image_filenames,
            }

        # 判断是否有变化（merge 策略：非空字段覆盖）
        changed_fields = []
        for key, new_val in data.items():
            if key == "name":
                continue  # name 是主键，不比较
            if new_val is None or new_val == "":
                continue  # 空值不覆盖
            old_val = getattr(existing, key, None)
            if new_val != old_val:
                changed_fields.append(key)

        # 检查专业是否变化
        existing_programs = await prog_repo.list_programs(
            self.session, existing.id
        )
        existing_prog_set = {
            (p.name, p.discipline_id) for p in existing_programs
        }
        new_prog_set = set()
        for p in programs:
            cat_name = p.get("category_name")
            disc_name = p.get("discipline_name")
            if cat_name and disc_name:
                cat = await disc_repo.get_category_by_name(
                    self.session, cat_name
                )
                if cat:
                    disc = await disc_repo.get_discipline_by_name(
                        self.session, cat.id, disc_name
                    )
                    if disc:
                        new_prog_set.add((p["program_name"], disc.id))
        if new_prog_set != existing_prog_set:
            changed_fields.append("programs")

        # 检查图片是否变化
        if logo_filename and logo_filename in image_data_map:
            # 简化：如果提供了新 logo 文件名，视为变化
            changed_fields.append("logo")
        if image_filenames:
            for fn in image_filenames:
                if fn in image_data_map:
                    changed_fields.append("images")
                    break

        status = "update" if changed_fields else "unchanged"

        return {
            "row": row_num,
            "name": name,
            "status": status,
            "changed_fields": changed_fields,
            "data": data,
            "programs": programs,
            "logo_filename": logo_filename,
            "image_filenames": image_filenames,
        }

    async def _find_unknown_disciplines(
        self, disc_paths: set[str]
    ) -> list[str]:
        """查找系统中不存在的学科分类。"""
        unknown = []
        for path in disc_paths:
            parts = path.split("/", 1)
            if len(parts) != 2:
                unknown.append(path)
                continue
            cat_name, disc_name = parts
            cat = await disc_repo.get_category_by_name(
                self.session, cat_name
            )
            if not cat:
                unknown.append(path)
                continue
            disc = await disc_repo.get_discipline_by_name(
                self.session, cat.id, disc_name
            )
            if not disc:
                unknown.append(path)
        return unknown

    async def _apply_discipline_actions(
        self, actions: list[dict]
    ) -> None:
        """根据用户决策创建或映射学科。

        actions: [{"name": "工学/量子计算", "action": "create"}]
        """
        for m in actions:
            if m.get("action") == "create":
                parts = m["name"].split("/", 1)
                if len(parts) != 2:
                    continue
                cat_name, disc_name = parts
                cat = await disc_repo.get_category_by_name(
                    self.session, cat_name
                )
                if not cat:
                    cat = DisciplineCategory(name=cat_name)
                    cat = await disc_repo.create_category(self.session, cat)
                disc = await disc_repo.get_discipline_by_name(
                    self.session, cat.id, disc_name
                )
                if not disc:
                    disc = Discipline(category_id=cat.id, name=disc_name)
                    await disc_repo.create_discipline(self.session, disc)

    async def _create_university(
        self, item: dict, image_data_map: dict[str, bytes]
    ) -> University:
        """创建新院校。"""
        data = item["data"]
        university = University(
            name=data["name"],
            name_en=data.get("name_en"),
            country=data["country"],
            province=data.get("province"),
            city=data["city"],
            website=data.get("website"),
            description=data.get("description"),
            admission_requirements=data.get("admission_requirements"),
            scholarship_info=data.get("scholarship_info"),
            latitude=data.get("latitude"),
            longitude=data.get("longitude"),
            is_featured=data.get("is_featured", False),
            sort_order=data.get("sort_order", 0),
            qs_rankings=data.get("qs_rankings"),
        )
        university = await uni_repo.create_university(
            self.session, university
        )

        # 上传 logo
        logo_filename = item.get("logo_filename")
        if logo_filename:
            logo_path = f"images/{logo_filename}"
            if logo_path in image_data_map:
                image = await image_repo.create_image(
                    self.session,
                    image_data_map[logo_path],
                    logo_filename,
                    self._guess_mime_type(logo_filename),
                )
                university.logo_image_id = image.id
                await uni_repo.update_university(self.session, university)

        # 上传展示图
        image_filenames = item.get("image_filenames", [])
        for idx, fn in enumerate(image_filenames[:5]):  # 最多 5 张
            img_path = f"images/{fn}"
            if img_path in image_data_map:
                image = await image_repo.create_image(
                    self.session,
                    image_data_map[img_path],
                    fn,
                    self._guess_mime_type(fn),
                )
                uni_image = UniversityImage(
                    university_id=university.id,
                    image_id=image.id,
                    sort_order=idx,
                )
                await uni_repo.add_university_image(self.session, uni_image)

        # 创建专业
        programs = []
        for p in item.get("programs", []):
            cat_name = p.get("category_name")
            disc_name = p.get("discipline_name")
            if cat_name and disc_name:
                cat = await disc_repo.get_category_by_name(
                    self.session, cat_name
                )
                if cat:
                    disc = await disc_repo.get_discipline_by_name(
                        self.session, cat.id, disc_name
                    )
                    if disc:
                        programs.append(
                            {
                                "name": p["program_name"],
                                "discipline_id": disc.id,
                            }
                        )
        if programs:
            await prog_repo.replace_programs(
                self.session, university.id, programs
            )

        return university

    async def _update_university(
        self, item: dict, image_data_map: dict[str, bytes]
    ) -> University:
        """更新院校（merge 策略）。"""
        name = item["name"]
        existing = await uni_repo.get_university_by_name(self.session, name)
        if not existing:
            raise ValueError(f"院校 {name} 不存在")

        # Merge 字段（非空覆盖）
        data = item["data"]
        for key, new_val in data.items():
            if key == "name":
                continue
            if new_val is None or new_val == "":
                continue  # 空值不覆盖
            setattr(existing, key, new_val)

        existing = await uni_repo.update_university(self.session, existing)

        # 更新 logo
        logo_filename = item.get("logo_filename")
        if logo_filename:
            logo_path = f"images/{logo_filename}"
            if logo_path in image_data_map:
                image = await image_repo.create_image(
                    self.session,
                    image_data_map[logo_path],
                    logo_filename,
                    self._guess_mime_type(logo_filename),
                )
                existing.logo_image_id = image.id
                await uni_repo.update_university(self.session, existing)

        # 更新展示图（全量替换）
        image_filenames = item.get("image_filenames", [])
        if image_filenames:
            # 删除旧图片记录
            old_images = await uni_repo.list_university_images(
                self.session, existing.id
            )
            for old_img in old_images:
                await uni_repo.delete_university_image(
                    self.session, old_img
                )
            # 添加新图片
            for idx, fn in enumerate(image_filenames[:5]):
                img_path = f"images/{fn}"
                if img_path in image_data_map:
                    image = await image_repo.create_image(
                        self.session,
                        image_data_map[img_path],
                        fn,
                        self._guess_mime_type(fn),
                    )
                    uni_image = UniversityImage(
                        university_id=existing.id,
                        image_id=image.id,
                        sort_order=idx,
                    )
                    await uni_repo.add_university_image(
                        self.session, uni_image
                    )

        # 更新专业（全量替换）
        programs = []
        for p in item.get("programs", []):
            cat_name = p.get("category_name")
            disc_name = p.get("discipline_name")
            if cat_name and disc_name:
                cat = await disc_repo.get_category_by_name(
                    self.session, cat_name
                )
                if cat:
                    disc = await disc_repo.get_discipline_by_name(
                        self.session, cat.id, disc_name
                    )
                    if disc:
                        programs.append(
                            {
                                "name": p["program_name"],
                                "discipline_id": disc.id,
                            }
                        )
        await prog_repo.replace_programs(
            self.session, existing.id, programs
        )

        return existing

    def _guess_mime_type(self, filename: str) -> str:
        """根据文件扩展名推测 MIME 类型。"""
        ext = filename.lower().split(".")[-1]
        mime_map = {
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "png": "image/png",
            "gif": "image/gif",
            "webp": "image/webp",
            "svg": "image/svg+xml",
        }
        return mime_map.get(ext, "image/jpeg")
