"""学科分类批量导入服务。"""

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.discipline import repository as disc_repo
from app.db.discipline.models import Discipline, DisciplineCategory
from app.utils.excel_io import (
    load_workbook_from_bytes,
    read_sheet_rows,
    workbook_to_bytes,
    write_sheet_header,
)


class DisciplineImportService:
    """学科分类导入服务。"""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def preview(self, content: bytes) -> dict:
        """解析 Excel，返回预览结果。"""
        wb = load_workbook_from_bytes(content)
        ws = wb.active
        rows = read_sheet_rows(ws)

        items = []
        errors = []

        for i, row in enumerate(rows, 2):
            if len(row) < 2 or not row[0] or not row[1]:
                errors.append({"row": i, "error": "大分类和小分类不能为空"})
                continue

            cat_name = str(row[0]).strip()
            disc_name = str(row[1]).strip()

            cat = await disc_repo.get_category_by_name(self.session, cat_name)
            disc = None
            if cat:
                disc = await disc_repo.get_discipline_by_name(
                    self.session, cat.id, disc_name
                )

            status = "unchanged" if disc else ("update" if cat else "new")
            items.append({
                "row": i,
                "category_name": cat_name,
                "discipline_name": disc_name,
                "status": status,
                "existing_category_id": cat.id if cat else None,
                "existing_discipline_id": disc.id if disc else None,
            })

        summary = {
            "new": sum(1 for it in items if it["status"] == "new"),
            "update": sum(1 for it in items if it["status"] == "update"),
            "unchanged": sum(1 for it in items if it["status"] == "unchanged"),
            "error": len(errors),
        }

        return {"items": items, "errors": errors, "summary": summary}

    async def confirm(self, items: list[dict]) -> dict:
        """执行导入。"""
        created = 0
        skipped = 0

        for item in items:
            if item["status"] == "unchanged":
                skipped += 1
                continue

            cat_name = item["category_name"]
            disc_name = item["discipline_name"]

            cat = await disc_repo.get_category_by_name(self.session, cat_name)
            if not cat:
                cat = DisciplineCategory(name=cat_name)
                cat = await disc_repo.create_category(self.session, cat)

            disc = await disc_repo.get_discipline_by_name(
                self.session, cat.id, disc_name
            )
            if not disc:
                disc = Discipline(category_id=cat.id, name=disc_name)
                await disc_repo.create_discipline(self.session, disc)
                created += 1
            else:
                skipped += 1

        return {"created": created, "skipped": skipped}

    def generate_template(self) -> bytes:
        """生成导入模板。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "学科分类"
        write_sheet_header(ws, ["大分类", "小分类"])
        ws.append(["工学", "计算机科学"])
        ws.append(["工学", "电子工程"])
        ws.append(["商学", "金融学"])
        ws.append(["医学", "临床医学"])
        return workbook_to_bytes(wb)
