"""成功案例批量导入服务。

支持 ZIP（包含 Excel + images/）和纯 Excel 两种格式。
Excel 格式：单个 sheet，表头：学生姓名 | 院校名称 | 专业 | 入学年份 | 感言 | 头像文件名 | Offer图文件名 | 是否精选 | 排序

导入策略：merge（非空字段覆盖，空字段保留）。
Merge key: (student_name, university, year)
"""

from fastapi import UploadFile
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import BadRequestException
from app.db.case import repository
from app.db.case.models import SuccessCase
from app.db.image import repository as image_repo
from app.db.university import repository as uni_repo
from app.utils.excel_io import (
    create_placeholder_image,
    create_zip,
    extract_zip,
    load_workbook_from_bytes,
    workbook_to_bytes,
    write_sheet_header,
)


class ImportService:
    """成功案例批量导入服务。"""

    def __init__(self, session: AsyncSession) -> None:
        """初始化服务。"""
        self.session = session

    async def preview(self, file: UploadFile) -> dict:
        """解析上传文件，返回预览结果。

        返回格式：
        {
          "items": [{"row": 2, "student_name": "张三", "status": "new|update|unchanged", "changed_fields": [...], "data": {...}}],
          "errors": [{"row": 3, "error": "..."}],
          "summary": {"new": 1, "update": 2, "unchanged": 0, "error": 1},
          "available_images": ["images/张三_avatar.jpg"]
        }
        """
        content = await file.read()
        filename = file.filename or ""

        # 解析 ZIP 或 Excel
        if filename.endswith(".zip"):
            files = extract_zip(content)
            excel_data = None
            for name, data in files.items():
                if name.endswith(".xlsx") and not name.startswith("__"):
                    excel_data = data
                    break
            if not excel_data:
                raise BadRequestException(
                    message="ZIP 中未找到 Excel 文件",
                    code="NO_EXCEL_IN_ZIP",
                )
            wb = load_workbook_from_bytes(excel_data)
            available_images = [
                name for name in files.keys() if name.startswith("images/")
            ]
        else:
            wb = load_workbook_from_bytes(content)
            available_images = []

        # 解析表格
        items, errors = await self._parse_workbook(wb)

        summary = {
            "new": sum(1 for x in items if x["status"] == "new"),
            "update": sum(1 for x in items if x["status"] == "update"),
            "unchanged": sum(1 for x in items if x["status"] == "unchanged"),
            "error": len(errors),
        }

        return {
            "items": items,
            "errors": errors,
            "summary": summary,
            "available_images": available_images,
        }

    async def confirm(
        self,
        file: UploadFile,
        items: list[dict],
    ) -> dict:
        """执行批量导入。

        Args:
            file: 原始上传文件（重新解析获取图片）
            items: preview 返回的 items（只需 data + avatar_filename + offer_filename）

        Returns:
            {"imported": 1, "updated": 2, "skipped": 3}
        """
        # 重新解析文件获取图片数据
        content = await file.read()
        filename = file.filename or ""
        image_data_map = {}
        if filename.endswith(".zip"):
            files = extract_zip(content)
            image_data_map = {
                name: files[name]
                for name in files.keys()
                if name.startswith("images/")
            }

        imported = 0
        updated = 0
        skipped = 0

        for item in items:
            try:
                if item["status"] == "new":
                    await self._create_case(item, image_data_map)
                    imported += 1
                elif item["status"] == "update":
                    await self._update_case(item, image_data_map)
                    updated += 1
                else:  # unchanged
                    skipped += 1
            except Exception:
                skipped += 1

        return {"imported": imported, "updated": updated, "skipped": skipped}

    def generate_template(self) -> bytes:
        """生成 ZIP 导入模板（包含 Excel + 空 images/ 目录）。"""
        wb = Workbook()

        # 单个 sheet
        ws = wb.active
        ws.title = "成功案例"
        headers = [
            "学生姓名",
            "院校名称",
            "专业",
            "入学年份",
            "感言",
            "头像文件名",
            "Offer图文件名",
            "是否精选",
            "排序",
        ]
        write_sheet_header(ws, headers)
        # 示例行
        ws.append(
            [
                "张三",
                "哈佛大学",
                "计算机科学",
                2026,
                "感谢老师们的帮助，梦想成真！",
                "张三_avatar.jpg",
                "张三_offer.jpg",
                "是",
                0,
            ]
        )

        excel_bytes = workbook_to_bytes(wb)

        # 创建 ZIP
        files = {
            "cases.xlsx": excel_bytes,
            "images/张三_avatar.jpg": create_placeholder_image(200, 200, "Avatar", "#6366f1"),
            "images/张三_offer.jpg": create_placeholder_image(800, 450, "Offer"),
        }
        return create_zip(files)

    async def _parse_workbook(
        self, wb: Workbook
    ) -> tuple[list[dict], list[dict]]:
        """解析 workbook，返回 (items, errors)。"""
        items = []
        errors = []

        # 获取第一个 sheet
        ws = wb.active
        if not ws:
            errors.append({"row": 0, "error": "Excel 文件为空"})
            return items, errors

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        # 解析每一行
        for idx, row in enumerate(rows, start=2):
            if not row[0]:
                continue  # 跳过空行

            try:
                item = await self._parse_row(idx, row)
                items.append(item)
            except ValueError as e:
                errors.append({"row": idx, "error": str(e)})

        return items, errors

    async def _parse_row(self, row_num: int, row: tuple) -> dict:
        """解析单行为 item。

        返回格式：
        {
          "row": 2,
          "student_name": "张三",
          "status": "new|update|unchanged",
          "changed_fields": ["program", "testimonial"],
          "data": {"student_name": "张三", "university": "哈佛大学", ...},
          "avatar_filename": "张三_avatar.jpg",
          "offer_filename": "张三_offer.jpg"
        }
        """
        student_name = str(row[0]).strip() if row[0] else ""
        if not student_name:
            raise ValueError("学生姓名不能为空")

        university = str(row[1]).strip() if row[1] else ""
        if not university:
            raise ValueError(f"行 {row_num}: 院校名称不能为空")

        program = str(row[2]).strip() if row[2] else ""
        if not program:
            raise ValueError(f"行 {row_num}: 专业不能为空")

        year = None
        if row[3]:
            try:
                year = int(row[3])
            except ValueError:
                raise ValueError(f"行 {row_num}: 入学年份必须是整数")
        if not year:
            raise ValueError(f"行 {row_num}: 入学年份不能为空")

        testimonial = str(row[4]).strip() if row[4] else None

        avatar_filename = str(row[5]).strip() if row[5] else None
        offer_filename = str(row[6]).strip() if row[6] else None

        is_featured = False
        if row[7]:
            val = str(row[7]).strip().lower()
            is_featured = val in ["是", "true", "1", "yes"]

        sort_order = 0
        if row[8]:
            try:
                sort_order = int(row[8])
            except ValueError:
                pass

        # 查询院校 ID（可选）
        university_id = None
        uni_obj = await uni_repo.get_university_by_name(self.session, university)
        if uni_obj:
            university_id = uni_obj.id

        # 查询数据库判断 new/update/unchanged
        existing = await repository.find_case(
            self.session, student_name, university, year
        )

        data = {
            "student_name": student_name,
            "university": university,
            "program": program,
            "year": year,
            "testimonial": testimonial,
            "is_featured": is_featured,
            "sort_order": sort_order,
            "university_id": university_id,
        }

        if not existing:
            return {
                "row": row_num,
                "student_name": student_name,
                "status": "new",
                "changed_fields": [],
                "data": data,
                "avatar_filename": avatar_filename,
                "offer_filename": offer_filename,
            }

        # 判断是否有变化（merge 策略：非空字段覆盖）
        changed_fields = []
        for key, new_val in data.items():
            if key in ["student_name", "university", "year"]:
                continue  # 这些是 merge key，不比较
            if new_val is None or new_val == "":
                continue  # 空值不覆盖
            old_val = getattr(existing, key, None)
            if new_val != old_val:
                changed_fields.append(key)

        # 检查图片是否变化
        if avatar_filename:
            changed_fields.append("avatar")
        if offer_filename:
            changed_fields.append("offer")

        status = "update" if changed_fields else "unchanged"

        return {
            "row": row_num,
            "student_name": student_name,
            "status": status,
            "changed_fields": changed_fields,
            "data": data,
            "avatar_filename": avatar_filename,
            "offer_filename": offer_filename,
        }

    async def _create_case(
        self, item: dict, image_data_map: dict[str, bytes]
    ) -> SuccessCase:
        """创建新案例。"""
        data = item["data"]
        case = SuccessCase(
            student_name=data["student_name"],
            university=data["university"],
            program=data["program"],
            year=data["year"],
            testimonial=data.get("testimonial"),
            is_featured=data.get("is_featured", False),
            sort_order=data.get("sort_order", 0),
            university_id=data.get("university_id"),
        )
        case = await repository.create_case(self.session, case)

        # 上传头像
        avatar_filename = item.get("avatar_filename")
        if avatar_filename:
            avatar_path = f"images/{avatar_filename}"
            if avatar_path in image_data_map:
                image = await image_repo.create_image(
                    self.session,
                    image_data_map[avatar_path],
                    avatar_filename,
                    self._guess_mime_type(avatar_filename),
                )
                case.avatar_image_id = image.id
                await repository.update_case(self.session, case)

        # 上传 Offer 图
        offer_filename = item.get("offer_filename")
        if offer_filename:
            offer_path = f"images/{offer_filename}"
            if offer_path in image_data_map:
                image = await image_repo.create_image(
                    self.session,
                    image_data_map[offer_path],
                    offer_filename,
                    self._guess_mime_type(offer_filename),
                )
                case.offer_image_id = image.id
                await repository.update_case(self.session, case)

        return case

    async def _update_case(
        self, item: dict, image_data_map: dict[str, bytes]
    ) -> SuccessCase:
        """更新案例（merge 策略）。"""
        data = item["data"]
        existing = await repository.find_case(
            self.session,
            data["student_name"],
            data["university"],
            data["year"],
        )
        if not existing:
            raise ValueError(
                f"案例 {data['student_name']} - {data['university']} ({data['year']}) 不存在"
            )

        # Merge 字段（非空覆盖）
        for key, new_val in data.items():
            if key in ["student_name", "university", "year"]:
                continue  # merge key 不更新
            if new_val is None or new_val == "":
                continue  # 空值不覆盖
            setattr(existing, key, new_val)

        existing = await repository.update_case(self.session, existing)

        # 更新头像
        avatar_filename = item.get("avatar_filename")
        if avatar_filename:
            avatar_path = f"images/{avatar_filename}"
            if avatar_path in image_data_map:
                image = await image_repo.create_image(
                    self.session,
                    image_data_map[avatar_path],
                    avatar_filename,
                    self._guess_mime_type(avatar_filename),
                )
                existing.avatar_image_id = image.id
                await repository.update_case(self.session, existing)

        # 更新 Offer 图
        offer_filename = item.get("offer_filename")
        if offer_filename:
            offer_path = f"images/{offer_filename}"
            if offer_path in image_data_map:
                image = await image_repo.create_image(
                    self.session,
                    image_data_map[offer_path],
                    offer_filename,
                    self._guess_mime_type(offer_filename),
                )
                existing.offer_image_id = image.id
                await repository.update_case(self.session, existing)

        return existing

    def _guess_mime_type(self, filename: str) -> str:
        """根据文件扩展名推测 MIME 类型。"""
        ext = filename.lower().split(".")[-1]
        mime_map = {
            "jpg": "image/jpeg",
            "jpeg": "image/jpeg",
            "png": "image/png",
            "gif": "image/gif",
            "webp": "image/webp",
            "svg": "image/svg+xml",
        }
        return mime_map.get(ext, "image/jpeg")
