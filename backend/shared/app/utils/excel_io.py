"""Excel/ZIP 读写通用工具。"""

import io
import zipfile

from PIL import Image, ImageDraw

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def read_sheet_rows(ws: Worksheet, skip_header: bool = True) -> list[list]:
    """读取 sheet 所有行，返回二维列表。"""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows[1:] if skip_header else rows


def write_sheet_header(ws: Worksheet, headers: list[str]) -> None:
    """写入表头行。"""
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)


def workbook_to_bytes(wb: Workbook) -> bytes:
    """将 Workbook 序列化为 bytes。"""
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def load_workbook_from_bytes(data: bytes) -> Workbook:
    """从 bytes 加载 Workbook。"""
    return load_workbook(io.BytesIO(data))


def extract_zip(content: bytes) -> dict[str, bytes]:
    """解压 ZIP，返回 {文件名: 内容} 字典。忽略 macOS 隐藏文件。"""
    files = {}
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        for name in zf.namelist():
            if name.startswith("__") or name.startswith("."):
                continue
            files[name] = zf.read(name)
    return files


def create_placeholder_image(
    width: int = 800, height: int = 450,
    text: str = "placeholder", color: str = "#94a3b8",
) -> bytes:
    """生成纯色占位 PNG 图片。"""
    img = Image.new("RGB", (width, height), color)
    draw = ImageDraw.Draw(img)
    bbox = draw.textbbox((0, 0), text)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    draw.text(((width - tw) / 2, (height - th) / 2), text, fill="white")
    buf = io.BytesIO()
    img.save(buf, "PNG")
    return buf.getvalue()


def create_zip(files: dict[str, bytes]) -> bytes:
    """打包文件为 ZIP，返回 bytes。"""
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in files.items():
            zf.writestr(name, data)
    output.seek(0)
    return output.getvalue()
